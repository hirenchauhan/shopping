from django.shortcuts import render,redirect,get_object_or_404
from myadmin.models import *
from customer.models import *
from saler.models import *
from django.contrib.auth.models import User
from django.contrib import auth,messages

#Date Filter
from django.utils.dateparse import parse_date
import datetime

#Download Exel File 
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import FileResponse

#Download Pdf File
from django.http import FileResponse,HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch 
from reportlab.lib.pagesizes import letter,A3
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Frame, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from io import BytesIO
from reportlab.lib.enums import TA_CENTER



def layout(request):
    contaxt = {}
    return render(request,'myadmin/common/layout.html',contaxt)

def dashboard(request):
    contaxt = {}
    return render(request,'myadmin/dashboard.html',contaxt)

def common_form(request):
    contaxt = {}
    return render(request,'myadmin/common_form.html',contaxt)

def common_table(request):
    contaxt = {}
    return render(request,'myadmin/common_table.html',contaxt)

# Login Crud
def login(request):
    context = {}
    return render(request,'myadmin/login.html',context)

def login_check(request):
    myusername  = request.POST['username']
    mypassword  = request.POST['password']

    result = auth.authenticate(username=myusername,password=mypassword)
    if result is None:
        messages.success(request,"Invaild username and password😅")
        return redirect('/myadmin/login')   
    else:
        auth.login(request,result)
        return redirect('/myadmin/dashboard')
    
def logout(request):
    auth.logout(request)
    return redirect('/myadmin/login')

# State Crud
def add_state(request):
    context = {}
    return render(request,'myadmin/add_state.html',context)

def insert_state(request):
    mystatename =  request.POST['state_name']

    State.objects.create(state_name = mystatename)
    return(redirect('/myadmin/add_state'))

def all_state(request):
    result = State.objects.all()
    context = {'result' : result}
    return render(request,'myadmin/all_state.html',context)

def del_state(request,id):
    result = State.objects.get(pk=id)
    result.delete()
    return redirect('/myadmin/all_state')

def edit_state(request,id):
    result = State.objects.get(pk=id)
    context = {'result' : result}
    return render(request,'myadmin/edit_state.html',context)

def update_state(request,id):
    mystatename =  request.POST['state_name']

    data = {
        'state_name' : mystatename
    }
    State.objects.update_or_create(pk=id,defaults=data)
    return redirect('/myadmin/all_state')


# City Crud
def add_city(request):
    result = State.objects.all()
    context = {'result' : result}
    return render(request,'myadmin/add_city.html',context)

def all_city(request):
    result = City.objects.all()
    context = {'result' :result}
    return render(request,'myadmin/all_city.html',context)

def insert_city(request):
	# result=State.objects.all()

	mycity=request.POST['city_name']
	mysid=request.POST['state']

	City.objects.create(city_name=mycity,state_id=mysid)
	return redirect('/myadmin/add_city')

def del_city(request,id):
    result = City.objects.get(pk=id)
    result.delete()
    return redirect('/myadmin/all_city')

def edit_city(request,id):
	city = City.objects.get(pk=id)
	state = State.objects.all()
	context={'city':city,'state':state}
	return render(request,'myadmin/edit_city.html',context)

def update_city(request,id):
	result=State.objects.all()
	mycity=request.POST['city_name']
	mystate=request.POST['state']

	data={
			'city_name':mycity,
			'state_id':mystate,
	    }
	City.objects.update_or_create(pk=id,defaults=data)
	return redirect('/myadmin/all_city')


# Area Crud
def add_area(request):
    city = City.objects.all()
    state = State.objects.all()
    context = {'city' : city,'state':state}
    return render(request,'myadmin/add_area.html',context)

def insert_area(request):
     myarea = request.POST['area_name']
     mycity = request.POST['city']
     mystate = request.POST['state']

     Area.objects.create(area_name=myarea,city_id=mycity,state_id=mystate)
     return redirect('/myadmin/add_area')

def all_area(request):
    result = Area.objects.all()
    context = {'result' : result}
    return render(request,'myadmin/all_area.html',context)

def del_area(request,id):
     result= Area.objects.get(pk=id)
     result.delete()
     return redirect('/myadmin/all_area')

def edit_area(request,id):
     area = Area.objects.get(pk=id)
     city = City.objects.all()
     state = State.objects.all()

     context = {'area':area,'city':city,'state':state}
     return render(request,'myadmin/edit_area.html',context)

def update_area(request,id):
    result=City.objects.all()
    myarea = request.POST['area_name']
    mycid = request.POST['city']
    mysid = request.POST['state']

    data={
            'area_name':myarea,
            'city_id':mycid,
            'state_id':mysid

    }
    Area.objects.update_or_create(pk=id,defaults=data)
    return redirect('/myadmin/all_area')

#User Crud
def customer(request):
    result = Register.objects.all()
    context = {'result':result}
    return render(request,'myadmin/customer.html',context)

def view_customer(request,id):
    result = Register.objects.get(pk=id)
    context = {'result':result}
    return render(request,'myadmin/view_customer.html',context)

def verify(request,id):
    #  id = request.session['id']
    #  mystatus = request.POST['status']
    data = {
        'status':'Verify'
    }
    Register.objects.update_or_create(pk=id, defaults=data)
    context = {}
    return redirect('/myadmin/customer')

def search_customer(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    result = []
    
    if start_date and end_date:
        start_date = parse_date(start_date)
        end_date = parse_date(end_date)
        if start_date and end_date:
            result = Register.objects.filter(date__range=[start_date, end_date])
        else:
            result = Register.objects.none()  
    else:
        result = Register.objects.all()
    print(f"Start Date: {start_date}, End Date: {end_date}, result: {result}")

    return render(request, 'myadmin/customer.html', {'result': result, 'start_date': start_date, 'end_date': end_date})


def gen_excel_customer(request):
    # Get the start and end dates from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    result = []
    
    if start_date and end_date:
        try:
            # Convert strings to datetime objects
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
            
            # Ensure the end_date includes the entire day by adding one day
            end_date += datetime.timedelta(days=1)

            # Filter Register objects based on the 'date' field
            result = Register.objects.filter(date__range=[start_date, end_date])
        except ValueError as e:
            print(f"Date conversion error: {e}")
    else:
        print("Start date or end date not provided")
    
    # Create an Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    
    # Add headers
    headers = ['ID', 'First Name', 'Last Name', 'Username', 'Email', 'Contact', 'Date of Birth', 'Gender', 'Address', 'State', 'City', 'Area']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data to the sheet
    for row_num, register in enumerate(result, 2):
        ws.cell(row=row_num, column=1, value=register.user.id)
        ws.cell(row=row_num, column=2, value=register.user.first_name)
        ws.cell(row=row_num, column=3, value=register.user.last_name)
        ws.cell(row=row_num, column=4, value=register.user.username)
        ws.cell(row=row_num, column=5, value=register.user.email)
        ws.cell(row=row_num, column=6, value=register.contact)
        ws.cell(row=row_num, column=7, value=register.dob.strftime('%Y-%m-%d') if register.dob else "")
        ws.cell(row=row_num, column=8, value=register.gender)
        ws.cell(row=row_num, column=9, value=register.address)
        ws.cell(row=row_num, column=10, value=register.state.state_name)
        ws.cell(row=row_num, column=11, value=register.city.city_name)
        ws.cell(row=row_num, column=12, value=register.area.area_name)
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save the Excel file to a buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    return FileResponse(buf, as_attachment=True, filename='customers.xlsx')

def gen_pdf_customer(request):
    # Get the start and end dates from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    result = []
    
    if start_date and end_date:
        try:
            # Convert strings to datetime objects
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
            
            # Ensure the end_date includes the entire day by adding one day
            end_date += datetime.timedelta(days=1)

            # Filter Register objects based on the 'date' field
            result = Register.objects.filter(date__range=[start_date, end_date])
        except ValueError as e:
            print(f"Date conversion error: {e}")
    else:
        print("Start date or end date not provided")
    
    # Create a PDF buffer
    buf = BytesIO()
    
    # Create a SimpleDocTemplate for the PDF
    doc = SimpleDocTemplate(buf, pagesize=A3)
    
    # Create a list to hold the PDF elements
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    header_style = styles['Heading1']
    normal_style = styles['Normal']
    
    # Add a title
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
    'HeaderStyle',
    parent=styles['Title'],
    alignment=TA_CENTER,
    fontName='Helvetica-Bold',
    fontSize=14
)
    elements.append(Paragraph('Customer List', header_style))
    
    # Define table data
    headers = ['ID', 'Full Name', 'Username', 'Email', 'Contact', 'Date of Birth', 'Gender', 'Address', 'State', 'City', 'Area']
    data = [headers]
    
    for register in result:
        row = [
            register.user.id,
            register.user.first_name + " " + register.user.last_name,
            register.user.username,
            register.user.email,
            register.contact,
            register.dob.strftime('%Y-%m-%d') if register.dob else "",
            register.gender,
            register.address,
            register.state.state_name,
            register.city.city_name,
            register.area.area_name,
        ]
        data.append(row)
    
    # Create a table with the data
    table = Table(data)
    
    # Define the table style
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f2dfdb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table.setStyle(table_style)
    
    # Add the table to the elements
    elements.append(table)
    
    # Build the PDF
    doc.build(elements)
    
    # Set the buffer position to the beginning
    buf.seek(0)
    
    # Return the PDF as a response
    return HttpResponse(buf, content_type='application/pdf', headers={'Content-Disposition': 'attachment; filename="customers.pdf"'})


#Contact Crud
def all_inquery(request):
    result = Contact.objects.all()
    context = {'result':result}
    return render(request,'myadmin/all_inquery.html',context)

def del_inquery(request,id):
    result = Contact.objects.get(pk=id)
    result.delete()
    return redirect('/myadmin/all_inquery')

#Catagories Crud
def add_categories(request):
    context = {}
    return render(request,'myadmin/add_categories.html',context)

def store_categories(request):
    mycat_name =  request.POST['cat_name']

    #insert
    Categories.objects.create(cat_name=mycat_name)
    return redirect('/myadmin/add_categories')


def all_categories(request):
    result = Categories.objects.all()
    context = {'result' : result}
    return render(request,'myadmin/all_categories.html',context)

def delete_cat(request,id):
    result = Categories.objects.get(pk=id)
    result.delete()
    return redirect('/myadmin/all_categories')

def edit_cat(request,id):
    result = Categories.objects.get(pk=id)
    context={'result':result}
    return render(request,'myadmin/edit_cat.html',context)

def update_cat(request,id):
    mycat_name =  request.POST['cat_name']

    data = {
            'cat_name':mycat_name
    }

    Categories.objects.update_or_create(pk=id,defaults=data)
    return redirect('/myadmin/all_categories')

#SubCategories Crud
def add_subcategories(request):
    result=Categories.objects.all()
    context = {'cat':result}
    return render(request,'myadmin/add_subcategories.html',context)

def store_subcat(request):
    result= Categories.objects.all()
    mysubcat_name = request.POST['subcat_name']
    mycid =  request.POST['cid']
   

    #insert
    Subcategories.objects.create(subcat_name=mysubcat_name,categories_id=mycid)
    return redirect('/myadmin/add_subcategories')

def all_subcategories(request):
    result = Subcategories.objects.all()
    context = {'result' : result}
    return render(request,'myadmin/all_subcategories.html',context)

def delete_subcat(request,id):
    result = Subcategories.objects.get(pk=id)
    result.delete()
    return redirect('/myadmin/all_subcategories')

def edit_subcat(request,id):
    result = Subcategories.objects.get(pk=id)
    result1 = Categories.objects.all()
    context={'result':result,'cat':result1}
    return render(request,'myadmin/edit_subcat.html',context)

def update_subcat(request,id):
    result= Categories.objects.all()
    mysubcat_name = request.POST['subcat_name']
    mycid =  request.POST['cid']
    

    data={
            'subcat_name':mysubcat_name,
            'categories_id':mycid

    }

    Subcategories.objects.update_or_create(pk=id,defaults=data)
    return redirect('/myadmin/all_subcategories')

#Saler
def saler(request):
    saler = Saler_reg.objects.all()
    context = {'saler' : saler}
    return render(request,'myadmin/saler.html',context)

def detail_salers(request,id):
    result = Saler_reg.objects.get(pk=id)
    context = {'result':result}
    return render(request,'myadmin/detail_salers.html',context)

def saler_verify(request,id):
    data = {
          'status':'verify'
     }
    Saler_reg.objects.update_or_create(pk=id, defaults=data)
    context = {}
    return redirect('/myadmin/saler')

def delete_saler(request,id):
    result = Saler_reg.objects.get(pk=id)
    result.delete()
    return redirect('/myadmin/saler')

def search_saler(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    saler = []
    
    if start_date and end_date:
        start_date = parse_date(start_date)
        end_date = parse_date(end_date)
        if start_date and end_date:
            saler = Saler_reg.objects.filter(date__range=[start_date, end_date])
        else:
            saler = Saler_reg.objects.none()  # or handle invalid dates differently
    else:
        saler = Saler_reg.objects.all()
    print(f"Start Date: {start_date}, End Date: {end_date}, saler: {saler}")

    return render(request, 'myadmin/saler.html', {'saler': saler, 'start_date': start_date, 'end_date': end_date})

def gen_excel_saler(request):
    # Get the start and end dates from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    result = []
    
    if start_date and end_date:
        try:
            # Convert strings to datetime objects
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
            
            # Ensure the end_date includes the entire day by adding one day
            end_date += datetime.timedelta(days=1)

            # Filter Saler_reg objects based on the 'date' field
            result = Saler_reg.objects.filter(date__range=[start_date, end_date])
        except ValueError as e:
            print(f"Date conversion error: {e}")
    else:
        print("Start date or end date not provided")
    
    # Create an Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Salers"
    
    # Add headers
    headers = [
        'ID', 'First Name', 'Last Name', 'Username', 'Email', 'Shop Contact', 'Image', 
        'Shop Name', 'Shop Address', 'State', 'City', 'Area'
          
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data to the sheet
    for row_num, saler in enumerate(result, 2):
        ws.cell(row=row_num, column=1, value=saler.id)
        ws.cell(row=row_num, column=2, value=saler.user.first_name)
        ws.cell(row=row_num, column=3, value=saler.user.last_name)
        ws.cell(row=row_num, column=4, value=saler.user.username)
        ws.cell(row=row_num, column=5, value=saler.user.email)
        ws.cell(row=row_num, column=6, value=saler.shop_contact)
        ws.cell(row=row_num, column=7, value=saler.image)
        ws.cell(row=row_num, column=8, value=saler.shop_name)
        ws.cell(row=row_num, column=9, value=saler.shop_address)
        ws.cell(row=row_num, column=10, value=saler.state.state_name)
        ws.cell(row=row_num, column=11, value=saler.city.city_name)
        ws.cell(row=row_num, column=12, value=saler.area.area_name)
        
        
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save the Excel file to a buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    return FileResponse(buf, as_attachment=True, filename='salers.xlsx')

def gen_pdf_saler(request):
    # Get the start and end dates from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    result = []
    
    if start_date and end_date:
        try:
            # Convert strings to datetime objects
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
            
            # Ensure the end_date includes the entire day by adding one day
            end_date += datetime.timedelta(days=1)

            # Filter Saler_reg objects based on the 'date' field
            result = Saler_reg.objects.filter(date__range=[start_date, end_date])
        except ValueError as e:
            print(f"Date conversion error: {e}")
    else:
        print("Start date or end date not provided")
    
    # Create a PDF buffer
    buf = BytesIO()
    
    # Create a SimpleDocTemplate for the PDF
    doc = SimpleDocTemplate(buf, pagesize=A3)
    
    # Create a list to hold the PDF elements
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    header_style = styles['Heading1']
    normal_style = styles['Normal']
    
    # Add a title
    elements.append(Paragraph('Saler List', header_style))
    
    # Define table data
    headers = [
        'ID', 'Full Name', 'Email', 'State', 'City', 'Area',
        'Shop Name', 'Shop Address', 'Shop Contact', 'Username'
    ]
    data = [headers]
    
    for saler in result:
        row = [
            saler.id,
            saler.user.first_name  + " " +  saler.user.last_name,
            saler.user.email,
            saler.state.state_name,
            saler.city.city_name,
            saler.area.area_name,
            saler.shop_name,
            saler.shop_address,
            saler.shop_contact,
            saler.user.username
        ]
        data.append(row)
    
    # Create a table with the data
    table = Table(data)
    
    # Define the table style
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#987D9A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table.setStyle(table_style)
    
    # Add the table to the elements
    elements.append(table)
    
    # Build the PDF
    doc.build(elements)
    
    # Set the buffer position to the beginning
    buf.seek(0)
    
    # Return the PDF as a response
    return HttpResponse(buf, content_type='application/pdf', headers={'Content-Disposition': 'attachment; filename="salers.pdf"'})

#Order
def order(request):
    
    order = Order.objects.all()
    context = {'order':order}
    return render(request,'myadmin/order.html',context)

def order_detail(request,id):
    order = get_object_or_404(Order, pk=id)
    order_details = Order_details.objects.filter(order=order)
    billing = get_object_or_404(Billing, order=order)
    shipping = get_object_or_404(Shipping, order=order)

    context = {
        'order': order,
        'order_details': order_details,
        'billing': billing,
        'shipping': shipping,
    }
    return render(request,'myadmin/order_detail.html',context)

def search_orders(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    order = []
    
    if start_date and end_date:
        start_date = parse_date(start_date)
        end_date = parse_date(end_date)
        if start_date and end_date:
            order = Order.objects.filter(date__range=[start_date, end_date])
        else:
            order = Order.objects.none()  # or handle invalid dates differently
    else:
        order = Order.objects.all()
    print(f"Start Date: {start_date}, End Date: {end_date}, order: {order}")

    return render(request, 'myadmin/order.html', {'order': order, 'start_date': start_date, 'end_date': end_date})


def gen_excel_order(request):
    # Get the start and end dates from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    result = []
    
    if start_date and end_date:
        try:
            # Convert strings to datetime objects
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
            
            # Ensure the end_date includes the entire day by adding one day
            end_date += datetime.timedelta(days=1)

            # Filter Order objects based on the 'date' field
            result = Order.objects.filter(date__range=[start_date, end_date])
        except ValueError as e:
            print(f"Date conversion error: {e}")
    else:
        print("Start date or end date not provided")
    
    # Create an Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    
    # Add headers
    headers = ['ID', 'Date', 'Amount', 'Status', 'Payment Method', 'Saler Name', 'User name']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data to the sheet
    for row_num, order in enumerate(result, 2):
        ws.cell(row=row_num, column=1, value=order.id)
        ws.cell(row=row_num, column=2, value=order.date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=3, value=order.amount)
        ws.cell(row=row_num, column=4, value=order.status)
        ws.cell(row=row_num, column=5, value=order.pay_method)
        ws.cell(row=row_num, column=6, value=order.saler.user.first_name)
        ws.cell(row=row_num, column=7, value=order.user.first_name)
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save the Excel file to a buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    return FileResponse(buf, as_attachment=True, filename='orders.xlsx')

def gen_pdf_orders(request):
    # Get the start and end dates from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    result = []
    
    if start_date and end_date:
        try:
            # Convert strings to datetime objects
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
            
            # Ensure the end_date includes the entire day by adding one day
            end_date += datetime.timedelta(days=1)

            # Filter Order objects based on the 'date' field
            result = Order.objects.filter(date__range=[start_date, end_date])
        except ValueError as e:
            print(f"Date conversion error: {e}")
    else:
        print("Start date or end date not provided")
    
    # Create a PDF buffer
    buf = BytesIO()
    
    # Create a SimpleDocTemplate for the PDF
    doc = SimpleDocTemplate(buf, pagesize=A3)
    
    # Create a list to hold the PDF elements
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    header_style = styles['Heading1']
    normal_style = styles['Normal']
    
    # Add a title
    elements.append(Paragraph('Order List', header_style))
    
    # Define table data
    headers = ['ID', 'Date', 'Amount', 'Status', 'Payment Method', 'Saler Name', 'User Name']
    data = [headers]
    
    for order in result:
        row = [
            order.id,
            order.date.strftime('%Y-%m-%d'),
            order.amount,
            order.status,
            order.pay_method,
            order.saler.user.first_name,
            order.user.first_name,
        ]
        data.append(row)
    
    # Create a table with the data
    table = Table(data)
    
    # Define the table style
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#987D9A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table.setStyle(table_style)
    
    # Add the table to the elements
    elements.append(table)
    
    # Build the PDF
    doc.build(elements)
    
    # Set the buffer position to the beginning
    buf.seek(0)
    
    # Return the PDF as a response
    return HttpResponse(buf, content_type='application/pdf', headers={'Content-Disposition': 'attachment; filename="orders.pdf"'})